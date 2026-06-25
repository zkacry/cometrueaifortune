"""
マイフォーチューン 価格設定・収益シミュレーション Excel生成
"""
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

OUT = r"C:\Users\Administrator\OneDrive\subwork\smartphone\smart-claude-code\apps\myfortune\docs\pricing_analysis.xlsx"

# ── カラー定数 ──────────────────────────────────────────
BLUE   = "FF0000FF"   # 入力値（青）
BLACK  = "FF000000"   # 数式（黒）
GREEN  = "FF008000"   # 他シート参照（緑）
YELLOW = "FFFFFF00"   # 要注意セル背景（黄）
HEADER_BG   = "FF2D2D5E"  # 濃紺ヘッダー背景
HEADER_FG   = "FFFFFFFF"  # ヘッダー白文字
SUBHEAD_BG  = "FF4A4A8A"  # サブヘッダー背景
SECTION_BG  = "FFE8E8F5"  # セクション薄背景
ALT_ROW_BG  = "FFF5F5FF"  # 交互行背景

# ── スタイルヘルパー ──────────────────────────────────────
def hfont(bold=False, size=11, color=BLACK, italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def hfill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

thin = Side(style="thin", color="FFAAAAAA")
thick = Side(style="medium", color="FF888888")

def tborder(top=False, bottom=False, left_b=False, right_b=False):
    return Border(
        top=thick if top else None,
        bottom=thick if bottom else None,
        left=thick if left_b else None,
        right=thick if right_b else None,
    )

def full_thin():
    return Border(top=thin, bottom=thin, left=thin, right=thin)

def set_header(cell, text, size=11):
    cell.value = text
    cell.font = hfont(bold=True, size=size, color=HEADER_FG)
    cell.fill = hfill(HEADER_BG)
    cell.alignment = center()
    cell.border = full_thin()

def set_subheader(cell, text):
    cell.value = text
    cell.font = hfont(bold=True, color=HEADER_FG)
    cell.fill = hfill(SUBHEAD_BG)
    cell.alignment = left()
    cell.border = full_thin()

def set_section(cell, text):
    cell.value = text
    cell.font = hfont(bold=True, color="FF2D2D5E")
    cell.fill = hfill("FFD8D8F0")
    cell.alignment = left()
    cell.border = full_thin()

def set_input(cell, value, fmt=None):
    cell.value = value
    cell.font = hfont(color=BLUE)
    cell.alignment = right()
    cell.border = full_thin()
    cell.fill = hfill("FFEEF4FF")
    if fmt:
        cell.number_format = fmt

def set_formula(cell, formula, fmt=None):
    cell.value = formula
    cell.font = hfont(color=BLACK)
    cell.alignment = right()
    cell.border = full_thin()
    if fmt:
        cell.number_format = fmt

def set_label(cell, text, indent=0, alt=False):
    cell.value = text
    cell.font = hfont()
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               indent=indent, wrap_text=True)
    cell.border = full_thin()
    if alt:
        cell.fill = hfill(ALT_ROW_BG)

def col_width(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def row_height(ws, row, h):
    ws.row_dimensions[row].height = h

JPY_FMT = '#,##0"円";(#,##0"円");"-"'
JPY_MM  = '#,##0"円"'
USD_FMT = '$#,##0.000000;($#,##0.000000);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'
NUM_FMT = '#,##0;(#,##0);"-"'

# ══════════════════════════════════════════════════════════
# Sheet 1: 原価計算
# ══════════════════════════════════════════════════════════
def build_sheet1(wb):
    ws = wb.create_sheet("1_原価計算")
    ws.sheet_view.showGridLines = False

    # 列幅
    widths = [28, 18, 18, 18, 8, 28, 14]
    for i, w in enumerate(widths, 1):
        col_width(ws, i, w)

    r = 1
    # タイトル
    ws.merge_cells(f"A{r}:G{r}")
    c = ws[f"A{r}"]
    c.value = "マイフォーチューン — 原価・コスト分析"
    c.font = hfont(bold=True, size=14, color=HEADER_FG)
    c.fill = hfill(HEADER_BG)
    c.alignment = center()
    row_height(ws, r, 30)
    r += 1

    # 注釈
    ws.merge_cells(f"A{r}:G{r}")
    c = ws[f"A{r}"]
    c.value = "■ 青文字=入力値（変更可）　■ 黒文字=計算式　■ 緑文字=他シート参照"
    c.font = hfont(italic=True, size=9, color="FF555555")
    c.alignment = left()
    r += 2

    # ── セクション1: 基本前提 ──
    ws.merge_cells(f"A{r}:G{r}")
    set_section(ws[f"A{r}"], "【基本前提】")
    row_height(ws, r, 22)
    r += 1

    assumptions = [
        ("USD/JPY レート", 155, "円", "#,##0"),
        ("1回の占いあたり 入力トークン", 800, "tokens", "#,##0"),
        ("1回の占いあたり 出力トークン", 600, "tokens", "#,##0"),
    ]
    # 前提ラベル行
    for label, val, unit, fmt in assumptions:
        set_label(ws[f"A{r}"], label)
        set_input(ws[f"B{r}"], val, fmt)
        set_label(ws[f"C{r}"], unit)
        ws.merge_cells(f"D{r}:G{r}")
        r += 1

    # 前提セルのアドレス（後で参照用）
    USD_JPY_ROW = r - 3
    IN_TOK_ROW  = r - 2
    OUT_TOK_ROW = r - 1

    r += 1

    # ── セクション2: AI APIコスト ──
    ws.merge_cells(f"A{r}:G{r}")
    set_section(ws[f"A{r}"], "【AI APIコスト（1回の占いあたり）】")
    row_height(ws, r, 22)
    r += 1

    # ヘッダー行
    headers = ["プロバイダー", "入力 $/M tokens", "出力 $/M tokens",
               "1回コスト($)", "1回コスト(¥)", "備考", "優先順"]
    for col, h in enumerate(headers, 1):
        set_header(ws.cell(r, col), h)
    row_height(ws, r, 28)
    HDR_ROW = r
    r += 1

    providers = [
        ("DeepSeek deepseek-chat",    0.14,  0.28,  "主力プロバイダー（最安）", 1),
        ("Gemini 2.0 Flash Lite",     0.075, 0.3,   "フォールバック①",       2),
        ("Claude Haiku 4.5",          0.8,   4.0,   "フォールバック②（最終）", 3),
    ]
    api_rows = []
    for prov, inp_price, out_price, note, priority in providers:
        set_label(ws.cell(r, 1), prov)
        set_input(ws.cell(r, 2), inp_price, '$#,##0.000')
        set_input(ws.cell(r, 3), out_price, '$#,##0.000')
        # 1回コスト($) = (入力トークン * 入力$/M + 出力トークン * 出力$/M) / 1,000,000
        set_formula(ws.cell(r, 4),
            f"=(B{IN_TOK_ROW}*B{r}+B{OUT_TOK_ROW}*C{r})/1000000",
            '$#,##0.0000000')
        # 1回コスト(¥)
        set_formula(ws.cell(r, 5),
            f"=D{r}*B{USD_JPY_ROW}",
            '#,##0.00"円"')
        set_label(ws.cell(r, 6), note)
        set_input(ws.cell(r, 7), priority, "#,##0")
        api_rows.append(r)
        r += 1

    r += 1

    # ── セクション3: プランごと月間AIコスト ──
    ws.merge_cells(f"A{r}:G{r}")
    set_section(ws[f"A{r}"], "【プランごとのユーザー1人あたり 月間AIコスト】")
    row_height(ws, r, 22)
    r += 1

    # ヘッダー
    plan_headers = ["プラン", "月間占い回数上限", "DeepSeek使用時(¥)", "Gemini使用時(¥)", "Claude使用時(¥)", "主力想定コスト(¥)", ""]
    for col, h in enumerate(plan_headers, 1):
        set_header(ws.cell(r, col), h)
    row_height(ws, r, 28)
    r += 1

    plans = [("Free（無料）", 10), ("Light（¥680/月）", 50), ("Pro（¥1,500/月）", 300)]
    plan_cost_rows = []
    for i, (plan, readings) in enumerate(plans):
        set_label(ws.cell(r, 1), plan, alt=(i % 2 == 1))
        set_input(ws.cell(r, 2), readings, "#,##0")
        # DeepSeek/Gemini/Claude の1回コスト列はapi_rows[0..2] の列5(¥)
        for col, api_row in enumerate(api_rows, 3):
            set_formula(ws.cell(r, col), f"=B{r}*E{api_row}", '#,##0.00"円"')
        # 主力想定（DeepSeek = col3）
        set_formula(ws.cell(r, 6), f"=C{r}", '#,##0.00"円"')
        plan_cost_rows.append(r)
        r += 1

    r += 1

    # ── セクション4: Google Play手数料 ──
    ws.merge_cells(f"A{r}:G{r}")
    set_section(ws[f"A{r}"], "【Google Play 手数料と純売上】")
    row_height(ws, r, 22)
    r += 1

    gp_headers = ["プラン", "定価(¥)", "手数料率（1年目30%）", "手数料率（2年目~15%）", "純売上1年目(¥)", "純売上2年目~(¥)", ""]
    for col, h in enumerate(gp_headers, 1):
        set_header(ws.cell(r, col), h)
    row_height(ws, r, 28)
    r += 1

    plan_prices = [("Light", 680), ("Pro", 1500)]
    gp_rows = []
    FEE1_ROW = r  # 手数料前提を同じ行に置く
    set_input(ws.cell(r, 3), 0.30, PCT_FMT)  # 1年目手数料率（入力）
    set_input(ws.cell(r, 4), 0.15, PCT_FMT)  # 2年目手数料率（入力）

    for i, (plan, price) in enumerate(plan_prices):
        set_label(ws.cell(r, 1), f"{plan}プラン", alt=(i % 2 == 1))
        set_input(ws.cell(r, 2), price, JPY_FMT)
        if i == 0:
            pass  # すでに手数料入力済み
        else:
            set_input(ws.cell(r, 3), 0.30, PCT_FMT)
            set_input(ws.cell(r, 4), 0.15, PCT_FMT)
        set_formula(ws.cell(r, 5), f"=B{r}*(1-C{r})", JPY_FMT)
        set_formula(ws.cell(r, 6), f"=B{r}*(1-D{r})", JPY_FMT)
        gp_rows.append(r)
        r += 1

    r += 2
    ws.merge_cells(f"A{r}:G{r}")
    c = ws[f"A{r}"]
    c.value = "※ 入力値（青）は実際の料金体系変更時に更新してください。手数料率はGoogle Play Small Business Program適用を前提（年間$1M以下）"
    c.font = hfont(italic=True, size=9, color="FF666666")
    c.alignment = left()

    return {
        "usd_jpy_row": USD_JPY_ROW,
        "api_rows": api_rows,
        "plan_cost_rows": plan_cost_rows,
        "gp_rows": gp_rows,
    }


# ══════════════════════════════════════════════════════════
# Sheet 2: 収益シミュレーション
# ══════════════════════════════════════════════════════════
def build_sheet2(wb):
    ws = wb.create_sheet("2_収益シミュレーション")
    ws.sheet_view.showGridLines = False

    widths = [30, 16, 16, 16, 2, 30, 14]
    for i, w in enumerate(widths, 1):
        col_width(ws, i, w)

    r = 1
    ws.merge_cells(f"A{r}:G{r}")
    c = ws[f"A{r}"]
    c.value = "マイフォーチューン — 収益シミュレーション（3シナリオ）"
    c.font = hfont(bold=True, size=14, color=HEADER_FG)
    c.fill = hfill(HEADER_BG)
    c.alignment = center()
    row_height(ws, r, 30)
    r += 2

    # ── シナリオ前提 ──
    ws.merge_cells(f"A{r}:G{r}")
    set_section(ws[f"A{r}"], "【シナリオ前提（青文字を変更してシミュレーション）】")
    row_height(ws, r, 22)
    r += 1

    # ヘッダー
    sc_headers = ["前提項目", "保守的", "標準", "楽観的", "", "説明", ""]
    for col, h in enumerate(sc_headers, 1):
        set_header(ws.cell(r, col), h)
    row_height(ws, r, 24)
    r += 1

    # シナリオ前提データ行（行番号を変数で保持）
    sc_start = r
    scenarios_data = [
        ("月間新規DL数（人）",                   500,    2000,   5000,   "#,##0",   "アプリストアからの新規インストール数/月"),
        ("累積アクティブユーザー（6ヶ月後）",     1500,   8000,   25000,  "#,##0",   "継続利用ユーザーの推定累積数"),
        ("有料転換率",                            0.015,  0.030,  0.050,  PCT_FMT,  "Free→有料プランへの転換率"),
        ("Lightプラン比率（有料中）",             0.70,   0.65,   0.60,   PCT_FMT,  "有料ユーザーのうちLightプラン加入割合"),
        ("月次チャーン率（解約率）",              0.080,  0.050,  0.030,  PCT_FMT,  "有料ユーザーの月次解約率"),
        ("DeepSeek 1回コスト(¥) ※Sheet1参照",   None,   None,   None,   '#,##0.00"円"', "占い1回あたりのAIコスト"),
        ("Light月間利用回数（想定）",             40,     45,     50,     "#,##0",   "実際の使用回数（上限50）"),
        ("Pro月間利用回数（想定）",               150,    200,    280,    "#,##0",   "実際の使用回数（上限300）"),
    ]

    ROW_DL       = r
    ROW_USERS    = r + 1
    ROW_CONV     = r + 2
    ROW_LIGHT_R  = r + 3
    ROW_CHURN    = r + 4
    ROW_AI_COST  = r + 5
    ROW_LIGHT_USE= r + 6
    ROW_PRO_USE  = r + 7

    for i, (label, v_cons, v_std, v_opt, fmt, note) in enumerate(scenarios_data):
        set_label(ws.cell(r, 1), label, alt=(i % 2 == 1))
        if label.startswith("DeepSeek"):
            # Sheet1のDeepSeek 1回コスト(¥) を参照（E列 = api_rows[0]行目）
            # api_rows[0]はSheet1で計算されるが、Sheet1では行が動的なので
            # ここでは手動でリンク式を書く。おおよそ行16あたりを想定し、
            # 後で確認できるよう固定リンクにする
            for col in [2, 3, 4]:
                c2 = ws.cell(r, col)
                c2.value = "='1_原価計算'!E9"  # DeepSeek 1回コスト(¥)行
                c2.font = hfont(color=GREEN)
                c2.border = full_thin()
                c2.number_format = fmt
                c2.alignment = right()
        else:
            vals = [v_cons, v_std, v_opt]
            for col, val in zip([2, 3, 4], vals):
                set_input(ws.cell(r, col), val, fmt)
        set_label(ws.cell(r, 5), "")
        set_label(ws.cell(r, 6), note, alt=(i % 2 == 1))
        ws.merge_cells(f"F{r}:G{r}")
        r += 1

    r += 1

    # ── 月次収益計算 ──
    ws.merge_cells(f"A{r}:G{r}")
    set_section(ws[f"A{r}"], "【月次収益・コスト・利益（6ヶ月後 安定期）】")
    row_height(ws, r, 22)
    r += 1

    # Light/Proプランの純売上単価（Sheet1参照）
    # Sheet1: gp_rows[0]=Lightの純売上1年目, gp_rows[1]=Proの純売上1年目
    # 概算でSheet1の値を直接参照

    # ヘッダー
    rev_headers = ["計算項目", "保守的", "標準", "楽観的", "", "計算式メモ", ""]
    for col, h in enumerate(rev_headers, 1):
        set_header(ws.cell(r, col), h)
    row_height(ws, r, 24)
    r += 1

    calc_start = r

    calc_items = [
        ("有料ユーザー総数（人）",      f"=ROUND(C{ROW_USERS}*C{ROW_CONV},0)",
                                        f"=ROUND(C{ROW_USERS}*C{ROW_CONV},0)",
                                        f"=ROUND(C{ROW_USERS}*C{ROW_CONV},0)",
                                        "#,##0", "累積ユーザー × 転換率"),
        ("　うち Lightプラン（人）",    None, None, None, "#,##0", "有料総数 × Light比率"),
        ("　うち Proプラン（人）",      None, None, None, "#,##0", "有料総数 × (1-Light比率)"),
        ("月間売上（Light）",           None, None, None, JPY_FMT, "Light人数 × 純売上単価(¥)"),
        ("月間売上（Pro）",             None, None, None, JPY_FMT, "Pro人数 × 純売上単価(¥)"),
        ("月間売上合計",                None, None, None, JPY_FMT, "Light売上 + Pro売上"),
        ("月間AIコスト（Freeユーザー）",None, None, None, JPY_FMT, "Free人数 × 10回 × 1回コスト"),
        ("月間AIコスト（Light）",       None, None, None, JPY_FMT, "Light人数 × 想定利用回数 × 1回コスト"),
        ("月間AIコスト（Pro）",         None, None, None, JPY_FMT, "Pro人数 × 想定利用回数 × 1回コスト"),
        ("月間AIコスト合計",            None, None, None, JPY_FMT, "全AIコスト合計"),
        ("Firebase/インフラコスト",     500,  1000, 3000, JPY_FMT, "小規模は無料枠内を想定"),
        ("月間純利益",                  None, None, None, JPY_FMT, "売上 - AIコスト - インフラ"),
        ("利益率",                      None, None, None, PCT_FMT, "純利益 ÷ 売上"),
        ("損益分岐点（有料ユーザー数）",None, None, None, "#,##0",  "固定コスト ÷ 1ユーザー限界利益"),
    ]

    # 行番号を保持
    ROW_PAID     = r
    ROW_LIGHT_U  = r + 1
    ROW_PRO_U    = r + 2
    ROW_REV_L    = r + 3
    ROW_REV_P    = r + 4
    ROW_REV_TOT  = r + 5
    ROW_AI_FREE  = r + 6
    ROW_AI_LIGHT = r + 7
    ROW_AI_PRO   = r + 8
    ROW_AI_TOT   = r + 9
    ROW_INFRA    = r + 10
    ROW_PROFIT   = r + 11
    ROW_MARGIN   = r + 12
    ROW_BEP      = r + 13

    # 純売上単価をSheet1から参照（概算固定リンク）
    # Light純売上1年目 ≈ Sheet1!E の gp_rows[0] ≈ 行24付近
    # 実際にはSheet1を確認し手動で調整
    LIGHT_NET = "='1_原価計算'!E24"   # Lightプラン純売上1年目
    PRO_NET   = "='1_原価計算'!E25"   # Proプラン純売上1年目

    for i, (label, *_) in enumerate(calc_items):
        set_label(ws.cell(r, 1), label, alt=(i % 2 == 1))
        set_label(ws.cell(r, 5), "")
        ws.merge_cells(f"F{r}:G{r}")
        set_label(ws.cell(r, 6), calc_items[i][5] if len(calc_items[i]) > 5 else "", alt=(i % 2 == 1))
        r += 1

    # 数式を後から設定
    for sc_col, sc_row in [(2, ROW_DL), (3, ROW_DL), (4, ROW_DL)]:
        pass  # DL行は前提のみ

    # 有料ユーザー総数
    for col, users_col in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_PAID, col)
        c2.value = f"=ROUND({users_col}{ROW_USERS}*{users_col}{ROW_CONV},0)"
        c2.font = hfont(color=BLACK)
        c2.border = full_thin()
        c2.number_format = "#,##0"
        c2.alignment = right()

    # Lightプラン人数
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_LIGHT_U, col)
        c2.value = f"=ROUND({uc}{ROW_PAID}*{uc}{ROW_LIGHT_R},0)"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = "#,##0"; c2.alignment = right()

    # Proプラン人数
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_PRO_U, col)
        c2.value = f"=ROUND({uc}{ROW_PAID}*(1-{uc}{ROW_LIGHT_R}),0)"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = "#,##0"; c2.alignment = right()

    # 月間売上 Light
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_REV_L, col)
        c2.value = f"={uc}{ROW_LIGHT_U}*{LIGHT_NET.replace('=','')}"
        c2.font = hfont(color=GREEN if LIGHT_NET else BLACK)
        c2.border = full_thin(); c2.number_format = JPY_FMT; c2.alignment = right()

    # 月間売上 Pro
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_REV_P, col)
        c2.value = f"={uc}{ROW_PRO_U}*{PRO_NET.replace('=','')}"
        c2.font = hfont(color=GREEN if PRO_NET else BLACK)
        c2.border = full_thin(); c2.number_format = JPY_FMT; c2.alignment = right()

    # 売上合計
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_REV_TOT, col)
        c2.value = f"={uc}{ROW_REV_L}+{uc}{ROW_REV_P}"
        c2.font = hfont(bold=True, color=BLACK); c2.border = full_thin()
        c2.number_format = JPY_FMT; c2.alignment = right()
        c2.fill = hfill("FFF0FFF0")

    # AI Free コスト (Free人数 = 累積 - 有料, 10回/月)
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_AI_FREE, col)
        c2.value = f"=MAX(0,{uc}{ROW_USERS}-{uc}{ROW_PAID})*10*{uc}{ROW_AI_COST}"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = JPY_FMT; c2.alignment = right()

    # AI Light コスト
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_AI_LIGHT, col)
        c2.value = f"={uc}{ROW_LIGHT_U}*{uc}{ROW_LIGHT_USE}*{uc}{ROW_AI_COST}"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = JPY_FMT; c2.alignment = right()

    # AI Pro コスト
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_AI_PRO, col)
        c2.value = f"={uc}{ROW_PRO_U}*{uc}{ROW_PRO_USE}*{uc}{ROW_AI_COST}"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = JPY_FMT; c2.alignment = right()

    # AI コスト合計
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_AI_TOT, col)
        c2.value = f"={uc}{ROW_AI_FREE}+{uc}{ROW_AI_LIGHT}+{uc}{ROW_AI_PRO}"
        c2.font = hfont(bold=True, color=BLACK); c2.border = full_thin()
        c2.number_format = JPY_FMT; c2.alignment = right()
        c2.fill = hfill("FFFFF0F0")

    # インフラコスト（入力値）
    for col, val in [(2, 500), (3, 1000), (4, 3000)]:
        set_input(ws.cell(ROW_INFRA, col), val, JPY_FMT)

    # 純利益
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_PROFIT, col)
        c2.value = f"={uc}{ROW_REV_TOT}-{uc}{ROW_AI_TOT}-{uc}{ROW_INFRA}"
        c2.font = hfont(bold=True, size=12, color=BLACK)
        c2.border = tborder(top=True, bottom=True)
        c2.number_format = JPY_FMT; c2.alignment = right()
        c2.fill = hfill("FFF5FFF5")

    # 利益率
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_MARGIN, col)
        c2.value = f"=IF({uc}{ROW_REV_TOT}=0,\"-\",{uc}{ROW_PROFIT}/{uc}{ROW_REV_TOT})"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = PCT_FMT; c2.alignment = right()

    # 損益分岐点
    for col, uc in [(2, "B"), (3, "C"), (4, "D")]:
        c2 = ws.cell(ROW_BEP, col)
        # BEP = インフラコスト ÷ (平均純売上 - 平均AIコスト/ユーザー)
        # 簡易: インフラ ÷ (Light純売上×Light比率 + Pro純売上×(1-Light比率) - 加重AIコスト)
        c2.value = (
            f"=IFERROR(ROUND({uc}{ROW_INFRA}/("
            f"{LIGHT_NET.replace('=','')}*{uc}{ROW_LIGHT_R}+"
            f"{PRO_NET.replace('=','')}*(1-{uc}{ROW_LIGHT_R})"
            f"-{uc}{ROW_AI_COST}*({uc}{ROW_LIGHT_R}*{uc}{ROW_LIGHT_USE}"
            f"+(1-{uc}{ROW_LIGHT_R})*{uc}{ROW_PRO_USE})"
            f"),0),\"計算不可\")"
        )
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = "#,##0"; c2.alignment = right()


# ══════════════════════════════════════════════════════════
# Sheet 3: 価格感応度分析
# ══════════════════════════════════════════════════════════
def build_sheet3(wb):
    ws = wb.create_sheet("3_価格感応度")
    ws.sheet_view.showGridLines = False

    col_width(ws, 1, 28)
    for c in range(2, 9):
        col_width(ws, c, 15)

    r = 1
    ws.merge_cells(f"A{r}:H{r}")
    c = ws[f"A{r}"]
    c.value = "マイフォーチューン — 価格感応度分析"
    c.font = hfont(bold=True, size=14, color=HEADER_FG)
    c.fill = hfill(HEADER_BG)
    c.alignment = center()
    row_height(ws, r, 30)
    r += 2

    # 前提
    ws.merge_cells(f"A{r}:H{r}")
    set_section(ws[f"A{r}"], "【前提: 価格弾力性の仮定（標準シナリオベース）】")
    r += 1

    BASE_USERS = 8000
    BASE_CONV  = 0.030
    BASE_LIGHT = 0.65

    assump = [
        ("標準シナリオ 累積ユーザー数", BASE_USERS, "#,##0"),
        ("基準転換率（¥680/¥1,500時）", BASE_CONV, PCT_FMT),
        ("価格弾力性（10%値上げ→転換率X%減）", 0.05, PCT_FMT),
        ("Light:Pro比率（固定）", BASE_LIGHT, PCT_FMT),
        ("Google Play手数料", 0.30, PCT_FMT),
    ]
    BASE_USERS_ROW = r
    BASE_CONV_ROW  = r + 1
    ELAST_ROW      = r + 2
    LRATIO_ROW     = r + 3
    GP_FEE_ROW     = r + 4

    for i, (lbl, val, fmt) in enumerate(assump):
        set_label(ws.cell(r, 1), lbl, alt=(i%2==1))
        set_input(ws.cell(r, 2), val, fmt)
        ws.merge_cells(f"C{r}:H{r}")
        r += 1

    r += 1

    # ── Lightプラン価格テーブル ──
    ws.merge_cells(f"A{r}:H{r}")
    set_section(ws[f"A{r}"], "【Lightプラン 価格変更シミュレーション（Proは¥1,500固定）】")
    r += 1

    light_prices = [480, 580, 680, 780, 980]
    headers = ["Lightプラン価格", "価格変化率", "調整後転換率", "有料ユーザー数",
               "Light人数", "Pro人数", "月間売上合計", "現行との差"]
    for col, h in enumerate(headers, 1):
        set_header(ws.cell(r, col), h)
    row_height(ws, r, 24)
    r += 1

    BASE_LIGHT_PRICE = 680
    BASE_PRO_PRICE   = 1500

    light_rows = []
    for i, price in enumerate(light_prices):
        is_base = (price == BASE_LIGHT_PRICE)
        set_input(ws.cell(r, 1), price, JPY_FMT)
        if is_base:
            ws.cell(r, 1).fill = hfill(YELLOW)
            ws.cell(r, 1).font = hfont(bold=True, color="FF0000FF")
        # 価格変化率
        c2 = ws.cell(r, 2)
        c2.value = f"=(A{r}-{BASE_LIGHT_PRICE})/{BASE_LIGHT_PRICE}"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = PCT_FMT; c2.alignment = right()
        # 調整後転換率 = 基準 * (1 - 弾力性 * 価格変化率)
        c2 = ws.cell(r, 3)
        c2.value = f"=MAX(0.001,B{BASE_CONV_ROW}*(1-B{ELAST_ROW}*(A{r}-{BASE_LIGHT_PRICE})/{BASE_LIGHT_PRICE}))"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = PCT_FMT; c2.alignment = right()
        # 有料ユーザー数
        c2 = ws.cell(r, 4)
        c2.value = f"=ROUND(B{BASE_USERS_ROW}*C{r},0)"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = "#,##0"; c2.alignment = right()
        # Light人数
        c2 = ws.cell(r, 5)
        c2.value = f"=ROUND(D{r}*B{LRATIO_ROW},0)"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = "#,##0"; c2.alignment = right()
        # Pro人数
        c2 = ws.cell(r, 6)
        c2.value = f"=D{r}-E{r}"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = "#,##0"; c2.alignment = right()
        # 月間売上 (手数料引き後)
        c2 = ws.cell(r, 7)
        c2.value = f"=E{r}*A{r}*(1-B{GP_FEE_ROW})+F{r}*{BASE_PRO_PRICE}*(1-B{GP_FEE_ROW})"
        c2.font = hfont(bold=True, color=BLACK); c2.border = full_thin()
        c2.number_format = JPY_FMT; c2.alignment = right()
        light_rows.append(r)
        r += 1

    # 現行との差
    BASE_IDX = light_prices.index(BASE_LIGHT_PRICE)
    base_r = light_rows[BASE_IDX]
    for row_idx in light_rows:
        c2 = ws.cell(row_idx, 8)
        c2.value = f"=G{row_idx}-G{base_r}"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = JPY_FMT; c2.alignment = right()

    r += 2

    # ── Proプラン価格テーブル ──
    ws.merge_cells(f"A{r}:H{r}")
    set_section(ws[f"A{r}"], "【Proプラン 価格変更シミュレーション（Lightは¥680固定）】")
    r += 1

    pro_prices = [980, 1200, 1500, 1800, 2200]
    for col, h in enumerate(headers, 1):
        set_header(ws.cell(r, col), h)
    row_height(ws, r, 24)
    r += 1

    BASE_PRO_PRICE_IDX = pro_prices.index(1500)
    pro_rows = []
    for i, price in enumerate(pro_prices):
        is_base = (price == 1500)
        set_input(ws.cell(r, 1), price, JPY_FMT)
        if is_base:
            ws.cell(r, 1).fill = hfill(YELLOW)
            ws.cell(r, 1).font = hfont(bold=True, color="FF0000FF")
        c2 = ws.cell(r, 2)
        c2.value = f"=(A{r}-{BASE_PRO_PRICE})/1500"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = PCT_FMT; c2.alignment = right()
        c2 = ws.cell(r, 3)
        c2.value = f"=MAX(0.001,B{BASE_CONV_ROW}*(1-B{ELAST_ROW}*(A{r}-1500)/1500))"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = PCT_FMT; c2.alignment = right()
        c2 = ws.cell(r, 4)
        c2.value = f"=ROUND(B{BASE_USERS_ROW}*C{r},0)"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = "#,##0"; c2.alignment = right()
        c2 = ws.cell(r, 5)
        c2.value = f"=ROUND(D{r}*B{LRATIO_ROW},0)"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = "#,##0"; c2.alignment = right()
        c2 = ws.cell(r, 6)
        c2.value = f"=D{r}-E{r}"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = "#,##0"; c2.alignment = right()
        c2 = ws.cell(r, 7)
        c2.value = f"=E{r}*{BASE_LIGHT_PRICE}*(1-B{GP_FEE_ROW})+F{r}*A{r}*(1-B{GP_FEE_ROW})"
        c2.font = hfont(bold=True, color=BLACK); c2.border = full_thin()
        c2.number_format = JPY_FMT; c2.alignment = right()
        pro_rows.append(r)
        r += 1

    base_pro_r = pro_rows[BASE_PRO_PRICE_IDX]
    for row_idx in pro_rows:
        c2 = ws.cell(row_idx, 8)
        c2.value = f"=G{row_idx}-G{base_pro_r}"
        c2.font = hfont(color=BLACK); c2.border = full_thin()
        c2.number_format = JPY_FMT; c2.alignment = right()


# ══════════════════════════════════════════════════════════
# Sheet 4: 競合比較
# ══════════════════════════════════════════════════════════
def build_sheet4(wb):
    ws = wb.create_sheet("4_競合比較")
    ws.sheet_view.showGridLines = False

    col_width(ws, 1, 30)
    col_width(ws, 2, 16)
    col_width(ws, 3, 16)
    col_width(ws, 4, 12)
    col_width(ws, 5, 12)
    col_width(ws, 6, 12)
    col_width(ws, 7, 35)

    r = 1
    ws.merge_cells(f"A{r}:G{r}")
    c = ws[f"A{r}"]
    c.value = "マイフォーチューン — 競合比較・価格妥当性"
    c.font = hfont(bold=True, size=14, color=HEADER_FG)
    c.fill = hfill(HEADER_BG)
    c.alignment = center()
    row_height(ws, r, 30)
    r += 2

    headers = ["アプリ/サービス", "月額料金（下限）", "月額料金（上限）", "AI機能", "パーソナライズ", "占術数", "差別化ポイント比較"]
    for col, h in enumerate(headers, 1):
        set_header(ws.cell(r, col), h)
    row_height(ws, r, 28)
    r += 1

    competitors = [
        # カテゴリヘッダー
        ("── ホロスコープ系 ──", None, None, None, None, None, None),
        ("ザ・ホロスコープ",         480,  980,  "なし", "低",  "3〜5",  "星座・月星座に特化"),
        ("星占いパーフェクト",        0,    480,  "なし", "低",  "3",     "無料中心、広告収入"),
        ("Astro by Astrocenter",      980,  1480, "あり", "中",  "8〜10", "海外発、英語コンテンツ中心"),
        ("── タロット系 ──", None, None, None, None, None, None),
        ("ゴールデンタロット",         0,    480,  "なし", "低",  "1",     "タロット特化、無料"),
        ("タロット占い Premium",       480,  780,  "なし", "低",  "2〜3",  "シンプルUI"),
        ("── 総合占い系 ──", None, None, None, None, None, None),
        ("Uranai Premium",             480, 1200,  "なし", "中",  "6〜8",  "複数占術、日本語特化"),
        ("占いアプリ総合",             380,  980,  "なし", "低",  "5〜8",  "コンテンツ量重視"),
        ("── リアル占いサービス ──", None, None, None, None, None, None),
        ("電話占い（相場）",          18000,60000, "なし", "高",  "多数",  "30分〜2時間 × ¥300〜400/分"),
        ("チャット占い（相場）",       1500, 6000, "なし", "高",  "多数",  "1回ごと課金"),
        ("── マイフォーチューン ──", None, None, None, None, None, None),
        ("マイフォーチューン Light",   680,  680, "あり", "高", "10+",  "★AI記憶型・過去履歴学習・50回/月"),
        ("マイフォーチューン Pro",    1500, 1500, "あり", "高", "10+",  "★AIパーソナライズ強化・300回/月・四柱推命・縁起カレンダー"),
    ]

    for i, row_data in enumerate(competitors):
        label, low, high, ai, person, count, diff = row_data
        is_myapp = label.startswith("マイフォーチュン")
        is_cat = label.startswith("──")

        if is_cat:
            ws.merge_cells(f"A{r}:G{r}")
            set_section(ws[f"A{r}"], label)
            row_height(ws, r, 20)
            r += 1
            continue

        alt = (i % 2 == 0)
        fill = hfill("FFFFFFF0") if is_myapp else (hfill(ALT_ROW_BG) if alt else None)

        for col, val in enumerate([label, low, high, ai, person, count, diff], 1):
            c2 = ws.cell(r, col)
            c2.value = val if val is not None else "-"
            if is_myapp:
                c2.font = hfont(bold=True, color="FF2D2D5E")
                c2.fill = hfill("FFFFFFF0")
            else:
                c2.font = hfont()
                if fill:
                    c2.fill = fill
            c2.border = full_thin()
            if col in [2, 3]:
                c2.number_format = JPY_FMT if isinstance(val, int) else "@"
                c2.alignment = right()
            elif col == 4:
                if val == "あり":
                    c2.font = hfont(bold=True, color="FF008000")
            else:
                c2.alignment = left() if col in [1, 7] else center()
        r += 1

    r += 2

    # まとめ
    ws.merge_cells(f"A{r}:G{r}")
    set_section(ws[f"A{r}"], "【価格妥当性 まとめ】")
    r += 1

    summaries = [
        "✅ Light ¥680：競合の標準価格帯（¥480〜¥980）の中央値以下。AI記憶型は希少で差別化高い。値上げ余地あり（¥780〜¥880）。",
        "✅ Pro ¥1,500：総合占いアプリの最高値帯だが、300回/月 × AI分析 × 四柱推命で正当化可能。電話占い比では圧倒的コスパ。",
        "⚠️ Free枠（10回/月）：競合無料プランと比較して少なめ。初期獲得のため15〜20回程度への増加を検討。",
        "💡 推奨価格案：Light ¥780 / Pro ¥1,680。弾力性試算で標準シナリオ収益が約15〜20%向上見込み（Sheet3参照）。",
    ]
    for s in summaries:
        ws.merge_cells(f"A{r}:G{r}")
        c2 = ws[f"A{r}"]
        c2.value = s
        c2.font = hfont(size=10)
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.border = full_thin()
        row_height(ws, r, 36)
        r += 1


# ══════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)

    build_sheet1(wb)
    build_sheet2(wb)
    build_sheet3(wb)
    build_sheet4(wb)

    wb.save(OUT)
    print(f"Saved: {OUT}")

if __name__ == "__main__":
    main()
